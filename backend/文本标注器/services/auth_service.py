"""认证服务模块 — 从 users.xlsx 读取账号，登录后签发 JWT。"""

import logging
import os
import secrets
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import jwt
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ExcelUser:
    """Excel 中的一条用户记录。"""

    id: int
    username: str
    password: str
    role: str = ""
    name: str = ""

# ============================================================
# JWT 密钥管理
# ============================================================

def _get_jwt_secret() -> str:
    """获取 JWT 签名密钥。

    优先级：
    1. 环境变量 JWT_SECRET
    2. 若未设置，随机生成并写入 backend/.env 文件，同时设置环境变量

    Returns:
        JWT 签名密钥字符串
    """
    secret = os.environ.get("JWT_SECRET")
    if secret:
        return secret

    # 生成安全的随机密钥
    secret = secrets.token_urlsafe(32)

    # 计算 backend/.env 路径
    # 当前文件: backend/文本标注器/services/auth_service.py
    # 需要定位到: backend/.env
    current_dir = os.path.dirname(os.path.abspath(__file__))        # services/
    parent_dir = os.path.dirname(current_dir)                        # 文本标注器/
    backend_dir = os.path.dirname(parent_dir)                        # backend/
    env_path = os.path.join(backend_dir, ".env")

    try:
        if os.path.exists(env_path):
            # 追加到已有 .env
            with open(env_path, "r", encoding="utf-8") as f:
                content = f.read()
            if "JWT_SECRET" not in content:
                with open(env_path, "a", encoding="utf-8") as f:
                    f.write(f"\n# 自动生成的 JWT 签名密钥\nJWT_SECRET={secret}\n")
                logger.info("JWT_SECRET 已追加到 %s", env_path)
        else:
            # 创建新的 .env
            with open(env_path, "w", encoding="utf-8") as f:
                f.write(f"# 自动生成的 JWT 签名密钥\nJWT_SECRET={secret}\n")
            logger.info(".env 文件已创建，JWT_SECRET 已写入 %s", env_path)
    except OSError as e:
        logger.warning("无法写入 .env 文件 (%s)，JWT_SECRET 仅保存在内存中", e)

    os.environ["JWT_SECRET"] = secret
    return secret


# 模块加载时获取密钥
JWT_SECRET = _get_jwt_secret()
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_DAYS = 7


# 账号 Excel 路径：backend/users.xlsx
USERS_XLSX_PATH = Path(__file__).resolve().parents[2] / "users.xlsx"


def _normalize_cell(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def load_excel_users() -> dict[str, ExcelUser]:
    """从 users.xlsx 读取账号，启动时加载一次。"""
    if not USERS_XLSX_PATH.exists():
        raise RuntimeError(f"用户账号文件不存在: {USERS_XLSX_PATH}")

    workbook = load_workbook(USERS_XLSX_PATH, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise RuntimeError("users.xlsx 为空")

    header = [_normalize_cell(cell).lower() for cell in rows[0]]
    required_headers = {"username", "password"}
    if not required_headers.issubset(set(header)):
        raise RuntimeError("users.xlsx 缺少 username/password 表头")

    username_index = header.index("username")
    password_index = header.index("password")
    role_index = header.index("role") if "role" in header else None
    name_index = header.index("name") if "name" in header else None

    users: dict[str, ExcelUser] = {}
    next_id = 1
    for row in rows[1:]:
        if row is None:
            continue

        username = _normalize_cell(row[username_index] if username_index < len(row) else None)
        password = _normalize_cell(row[password_index] if password_index < len(row) else None)
        if not username or not password:
            continue

        users[username] = ExcelUser(
            id=next_id,
            username=username,
            password=password,
            role=_normalize_cell(row[role_index] if role_index is not None and role_index < len(row) else None),
            name=_normalize_cell(row[name_index] if name_index is not None and name_index < len(row) else None),
        )
        next_id += 1

    if not users:
        raise RuntimeError("users.xlsx 中没有可用账号")

    logger.info("已从 %s 加载 %d 个账号", USERS_XLSX_PATH, len(users))
    return users


EXCEL_USERS_BY_USERNAME = load_excel_users()


# ============================================================
# JWT 令牌工具函数
# ============================================================

def create_token(user_id: int, username: str) -> str:
    """生成 JWT 访问令牌。

    Args:
        user_id: 用户 ID
        username: 用户名

    Returns:
        JWT 令牌字符串
    """
    payload = {
        "user_id": user_id,
        "username": username,
        "exp": datetime.utcnow() + timedelta(days=JWT_EXPIRATION_DAYS),
        "iat": datetime.utcnow(),
    }
    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return token


def verify_token(token: str) -> dict | None:
    """验证 JWT 令牌的有效性。

    Args:
        token: JWT 令牌字符串

    Returns:
        解析成功的 payload 字典，验证失败返回 None
    """
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        logger.info("JWT 令牌已过期")
        return None
    except jwt.InvalidTokenError as e:
        logger.info("JWT 令牌无效: %s", e)
        return None


# ============================================================
# 认证业务逻辑
# ============================================================

def get_user_by_username(username: str) -> ExcelUser | None:
    return EXCEL_USERS_BY_USERNAME.get(username)


def get_user_by_id(user_id: int) -> ExcelUser | None:
    for user in EXCEL_USERS_BY_USERNAME.values():
        if user.id == user_id:
            return user
    return None


def login_user(username: str, password: str) -> dict:
    """使用 users.xlsx 中的账号密码登录。"""
    user = get_user_by_username(username)
    if not user or user.password != password:
        raise ValueError("用户名或密码错误")

    token = create_token(user.id, user.username)
    logger.info("用户登录成功: %s (id=%d)", username, user.id)

    return {
        "token": token,
        "user": {
            "id": user.id,
            "username": user.username,
            "role": user.role,
            "name": user.name or user.username,
        },
    }
